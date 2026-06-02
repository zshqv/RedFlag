import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CATEGORIES = ["Financial", "Regulatory", "Operational", "Legal"]


def analyze_trend(ticker):
    """Fetch up to 5 years of 10-K filings and return per-year category risk counts."""
    try:
        from fetchers.edgar_fetcher import get_cik_from_ticker, get_10k_filings, download_10k_text
    except ImportError:
        from edgar_fetcher import get_cik_from_ticker, get_10k_filings, download_10k_text

    from text_parser import extract_sections
    from risk_analyzer import analyze_filings

    print(f"[RedFlag] Trend: fetching up to 5 years of 10-K filings for {ticker}...")

    cik = get_cik_from_ticker(ticker)
    if not cik:
        return {"error": f"Could not find CIK for ticker '{ticker}'"}

    filings = get_10k_filings(cik)
    if not filings:
        return {"error": f"No 10-K filings found for '{ticker}'"}

    # One filing per calendar year, up to 5
    seen_years = set()
    target_filings = []
    for f in filings:
        year = f["date"][:4]
        if year not in seen_years:
            seen_years.add(year)
            target_filings.append(f)
        if len(target_filings) >= 5:
            break

    results = {}  # int(year) -> summary dict

    for filing in target_filings:
        year = int(filing["date"][:4])
        try:
            print(f"[RedFlag] Trend: downloading {year} filing...")
            time.sleep(0.3)
            text = download_10k_text(filing)
            if not text:
                print(f"[RedFlag] Trend: skipping {year} -- download returned None")
                continue
            sections = extract_sections(text)
            analysis = analyze_filings(sections)
            results[year] = analysis.get("summary", {})
            print(f"[RedFlag] Trend: {year} done ({len(analysis.get('findings', []))} findings)")
        except Exception as e:
            print(f"[RedFlag] Trend: skipping {year} -- {e}")
            continue

    if len(results) < 2:
        return {"error": "Insufficient historical data (fewer than 2 years fetched successfully)"}

    years_sorted = sorted(results.keys())
    trend = {"years": years_sorted}
    for cat in CATEGORIES:
        trend[cat] = [results[y].get(cat, 0) for y in years_sorted]

    return trend


def add_trend_sheet(wb, trend_data):
    """Add a '5-Year Trend' sheet to an openpyxl workbook."""
    ws = wb.create_sheet("5-Year Trend")

    if "error" in trend_data:
        cell = ws["A1"]
        cell.value = str(trend_data["error"])
        cell.font  = Font(name="Calibri", italic=True, color="AA0000")
        return

    years = trend_data.get("years", [])

    NAVY = "1A1A2E"
    CAT_COLORS = {
        "Financial":   "2980B9",
        "Regulatory":  "D35400",
        "Operational": "16A085",
        "Legal":       "8E44AD",
    }

    # Header row: Category label + one column per year
    hdr = ws.cell(row=1, column=1)
    hdr.value     = "Category"
    hdr.font      = Font(name="Calibri", bold=True, color="FFFFFF")
    hdr.fill      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    hdr.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16

    for ci, year in enumerate(years, 2):
        cell = ws.cell(row=1, column=ci)
        cell.value     = str(year)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12

    # One row per category
    for ri, cat in enumerate(CATEGORIES, 2):
        cat_cell           = ws.cell(row=ri, column=1)
        cat_cell.value     = cat
        cat_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF")
        cat_cell.fill      = PatternFill(
            start_color=CAT_COLORS.get(cat, NAVY),
            end_color=CAT_COLORS.get(cat, NAVY),
            fill_type="solid"
        )
        cat_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[ri].height = 22

        counts    = trend_data.get(cat, [])
        max_count = max(counts) if counts else 1

        for ci, count in enumerate(counts, 2):
            cell           = ws.cell(row=ri, column=ci)
            cell.value     = count
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(name="Calibri", bold=True)
            if max_count > 0 and count > 0:
                ratio = count / max_count
                if ratio >= 0.7:
                    bg = "FF4444"
                elif ratio >= 0.4:
                    bg = "FF8C00"
                else:
                    bg = "D4EFDF"
            else:
                bg = "F5F5F5"
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    ws.row_dimensions[1].height = 20


if __name__ == "__main__":
    print("trend_analyzer imported successfully")
